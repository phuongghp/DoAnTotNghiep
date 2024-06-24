#%%
import os
import xlrd
import openpyxl

def convert_xls_to_xlsx(xls_file, xlsx_file):
    # Open the input .xls file.
    xls_workbook = xlrd.open_workbook(xls_file)

    # Create a new .xlsx workbook.
    xlsx_workbook = openpyxl.Workbook()

    # Copy each sheet from the input .xls file to the output .xlsx file.
    for sheet_name in xls_workbook.sheet_names():
        xls_sheet = xls_workbook.sheet_by_name(sheet_name)
        xlsx_sheet = xlsx_workbook.create_sheet(sheet_name)

        # Copy the data from the input sheet to the output sheet.
        for row in range(xls_sheet.nrows):
            for col in range(xls_sheet.ncols):
                xlsx_sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell(row, col).value
    del xlsx_workbook['Sheet']
    # Save the output .xlsx file.
    xlsx_workbook.save(xlsx_file)

def add_excel(source_path,dest_path):
    # source_path = '/content/output.xlsx'
    source_wb = openpyxl.load_workbook(filename=source_path)

    # Load the destination workbook
    # dest_path = '/content/output1.xlsx'
    dest_wb = openpyxl.load_workbook(filename=dest_path)

    # Copy each sheet from source to destination
    for sheet in source_wb:
        # Create a new sheet in the destination workbook with the same title
        dest_sheet = dest_wb.create_sheet(sheet.title)

        # Copy data from each cell of the source sheet to the destination sheet
        for row in sheet.iter_rows():
            for cell in row:
                dest_sheet[cell.coordinate].value = cell.value

    # Save the destination workbook
    dest_wb.save(dest_path)

def remove_sheet(filename, sheet_name):
  wb = openpyxl.load_workbook(filename)
  del wb[sheet_name]
  wb.save(filename)

folder_path= 'D:\DATN\Code3\K59'

path_ex = folder_path+'/excel_files/'
if not os.path.exists(path_ex):
    os.mkdir(path_ex)
name_xls = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]

for i in name_xls:
    convert_xls_to_xlsx(folder_path+'/'+i, folder_path+'/excel_files/'+i+'x')

if os.path.exists(folder_path+'/excel_files/DaTNK59.xlsx'):
    os.remove(folder_path+'/excel_files/DaTNK59.xlsx')
else:
    # Create the file if it does not exist.
    out_workbook = openpyxl.Workbook()
    # out_workbook.create_sheet('sheet_name')
    out_workbook.save(folder_path+'/excel_files/DaTNK59.xlsx')

for i in name_xls:
    add_excel(folder_path+'/excel_files/'+i+'x',folder_path+'/excel_files/DaTNK59.xlsx')

remove_sheet(folder_path+'/excel_files/DaTNK59.xlsx', 'Sheet')
print(f'DONE Save to {folder_path}/excel_files/DaTNK59.xlsx')